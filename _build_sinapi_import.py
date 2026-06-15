from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook


BASE_DIR = Path("Gestão de obras") / "SINAPI_fontes" / "SINAPI-2026-05-formato-xlsx"
DEFAULT_REF_XLSX = BASE_DIR / "SINAPI_Referência_2026_05.xlsx"
DEFAULT_OUT_ROOT = Path("Gestão de obras") / "SINAPI_importacao_previa"

UFS = ("MG", "RJ", "SP")
REGIMES = {
    "NAO_DESONERADO": {"composicoes": "CSD", "insumos": "ISD", "label": "SEM DESONERAÇÃO"},
    "DESONERADO": {"composicoes": "CCD", "insumos": "ICD", "label": "COM DESONERAÇÃO"},
}


MONEY = Decimal("0.0001")
UUID_NAMESPACE = uuid5(NAMESPACE_URL, "https://webersongoncalves-lgtm.github.io/custoobra/orcamento")


@dataclass(frozen=True)
class CompositionCost:
    grupo: str
    codigo: str
    descricao: str
    unidade: str
    custo: Decimal
    as_percent: Decimal | None


@dataclass(frozen=True)
class AnalyticItem:
    comp_codigo: str
    tipo_item: str
    item_codigo: str
    descricao: str
    unidade: str
    coeficiente: Decimal
    situacao: str
    linha: int
    ordem: int


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").strip()


def code(value) -> str:
    raw = text(value)
    if not raw:
        return ""
    if raw.endswith(".0") and raw[:-2].isdigit():
        return raw[:-2]
    if raw.startswith("=HYPERLINK"):
        match = re.search(r",\s*(\d+)\s*\)\s*$", raw)
        if match:
            return match.group(1)
    return raw


def number(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = text(value)
    if not raw or raw == "-":
        return None
    raw = raw.replace("R$", "").replace("%", "").strip()
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def money(value: Decimal | None) -> Decimal:
    if value is None:
        return Decimal("0.0000")
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def stable_id(kind: str, key: str) -> str:
    return str(uuid5(UUID_NAMESPACE, f"{kind}|{key}"))


def load_uf_price_columns(ws, header_row: int) -> dict[str, int]:
    columns = {}
    for col in range(1, ws.max_column + 1):
        uf = text(ws.cell(header_row, col).value)
        if uf in UFS:
            columns[uf] = col
    missing = [uf for uf in UFS if uf not in columns]
    if missing:
        raise RuntimeError(f"UFs nao encontradas em {ws.title}: {missing}")
    return columns


def load_uf_cost_columns(ws) -> dict[str, tuple[int, int]]:
    columns = {}
    for col in range(1, ws.max_column + 1):
        uf = text(ws.cell(9, col).value)
        if uf in UFS:
            columns[uf] = (col, col + 1)
    missing = [uf for uf in UFS if uf not in columns]
    if missing:
        raise RuntimeError(f"UFs nao encontradas em {ws.title}: {missing}")
    return columns


def load_insumo_prices(ws):
    uf_cols = load_uf_price_columns(ws, 10)
    prices = {uf: {} for uf in UFS}
    metadata = {}
    for row in ws.iter_rows(min_row=11, values_only=True):
        item_code = code(row[1])
        if not item_code:
            continue
        classificacao = text(row[0])
        descricao = text(row[2])
        unidade = text(row[3]) or "NA"
        origem_preco = text(row[4])
        metadata[item_code] = {
            "classificacao": classificacao,
            "descricao": descricao,
            "unidade": unidade,
            "origem_preco": origem_preco,
        }
        for uf, col in uf_cols.items():
            preco = number(row[col - 1])
            if preco is not None:
                prices[uf][item_code] = money(preco)
    return prices, metadata


def load_composition_costs(ws_values, ws_formulas) -> dict[str, dict[str, CompositionCost]]:
    uf_cols = load_uf_cost_columns(ws_values)
    costs = {uf: {} for uf in UFS}
    value_rows = ws_values.iter_rows(min_row=11, values_only=True)
    formula_rows = ws_formulas.iter_rows(min_row=11, values_only=True)
    for row, formula_row in zip(value_rows, formula_rows):
        comp_code = code(formula_row[1]) or code(row[1])
        if not comp_code:
            continue
        grupo = text(row[0])
        descricao = text(row[2])
        unidade = text(row[3]) or "NA"
        for uf, (cost_col, as_col) in uf_cols.items():
            custo = number(row[cost_col - 1])
            if custo is None or custo <= 0:
                continue
            costs[uf][comp_code] = CompositionCost(
                grupo=grupo,
                codigo=comp_code,
                descricao=descricao,
                unidade=unidade,
                custo=money(custo),
                as_percent=number(row[as_col - 1]),
            )
    return costs


def load_analytic(ws):
    headers = {}
    items = defaultdict(list)
    item_orders = defaultdict(int)
    for row_idx, row in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
        comp_code = code(row[1])
        if not comp_code:
            continue
        tipo_item = text(row[2]).upper()
        item_code = code(row[3])
        descricao = text(row[4])
        unidade = text(row[5]) or "NA"
        coef = number(row[6])
        situacao = text(row[7])
        grupo = text(row[0])

        if not tipo_item and not item_code:
            headers[comp_code] = {
                "grupo": grupo,
                "descricao": descricao,
                "unidade": unidade,
                "situacao": situacao,
                "linha": row_idx,
            }
            continue

        if tipo_item not in {"INSUMO", "COMPOSICAO"} or not item_code:
            continue

        item_orders[comp_code] += 1
        items[comp_code].append(
            AnalyticItem(
                comp_codigo=comp_code,
                tipo_item=tipo_item,
                item_codigo=item_code,
                descricao=descricao,
                unidade=unidade,
                coeficiente=coef or Decimal("0"),
                situacao=situacao,
                linha=row_idx,
                ordem=item_orders[comp_code],
            )
        )
    return headers, items


def item_type(tipo_item: str, classificacao: str) -> str:
    if tipo_item == "COMPOSICAO":
        return "SUB"
    cls = classificacao.upper()
    if "MAO DE OBRA" in cls or "ENCARGOS COMPLEMENTARES" in cls:
        return "MO"
    if "EQUIPAMENTO" in cls:
        return "EQP"
    if "MATERIAL" in cls:
        return "MAT"
    if "SERVI" in cls:
        return "SUB"
    return "OUTRO"


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main(xlsx_path: Path = DEFAULT_REF_XLSX, out_dir: Path | None = None, ufs: tuple[str, ...] = UFS) -> None:
    global UFS
    UFS = ufs
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    wb_formulas = load_workbook(xlsx_path, read_only=True, data_only=False)

    competencia_raw = text(wb["CSD"].cell(3, 2).value)
    mes, ano = competencia_raw.split("/")
    competencia = f"{ano}-{mes.zfill(2)}"
    final_out_dir = Path(out_dir) if out_dir else DEFAULT_OUT_ROOT / competencia
    final_out_dir.mkdir(parents=True, exist_ok=True)

    analytic_headers, analytic_items = load_analytic(wb["Analítico"])

    regime_costs = {}
    regime_insumo_prices = {}
    regime_insumo_meta = {}
    for regime, cfg in REGIMES.items():
        regime_costs[regime] = load_composition_costs(wb[cfg["composicoes"]], wb_formulas[cfg["composicoes"]])
        prices, meta = load_insumo_prices(wb[cfg["insumos"]])
        regime_insumo_prices[regime] = prices
        regime_insumo_meta[regime] = meta

    wb.close()
    wb_formulas.close()

    services = {}
    compositions = []
    items = []
    audit = []
    missing_prices = []

    for regime, cfg in REGIMES.items():
        for uf in UFS:
            for comp_code, comp in sorted(regime_costs[regime][uf].items()):
                header = analytic_headers.get(comp_code, {})
                service_key = f"SINAPI-{comp_code}"
                services.setdefault(
                    service_key,
                    {
                        "id": stable_id("servico", service_key),
                        "service_key": service_key,
                        "codigo_interno": service_key,
                        "codigo_sinapi": comp_code,
                        "descricao": header.get("descricao") or comp.descricao,
                        "unidade": header.get("unidade") or comp.unidade,
                        "etapa": header.get("grupo") or comp.grupo,
                        "disciplina": "SINAPI",
                        "situacao": header.get("situacao", ""),
                    },
                )

                comp_key = f"SINAPI|{competencia}|{regime}|{uf}|{comp_code}"
                comp_id = stable_id("composicao", comp_key)
                compositions.append(
                    {
                        "id": comp_id,
                        "composition_key": comp_key,
                        "servico_id": stable_id("servico", service_key),
                        "service_key": service_key,
                        "codigo": comp_code,
                        "descricao": comp.descricao,
                        "unidade": comp.unidade,
                        "uf": uf,
                        "competencia": competencia,
                        "regime": regime,
                        "origem": "oficial",
                        "custo_unitario": str(comp.custo),
                        "referencia_preco": f"SINAPI {competencia} {cfg['label']} {uf}",
                        "grupo": comp.grupo,
                        "as_percent": "" if comp.as_percent is None else str(comp.as_percent),
                    }
                )

                sum_items = Decimal("0")
                count_items = 0
                for item in analytic_items.get(comp_code, []):
                    preco = Decimal("0")
                    classificacao = ""
                    origem_preco = ""
                    preco_fallback_uf = ""
                    if item.tipo_item == "INSUMO":
                        meta = regime_insumo_meta[regime].get(item.item_codigo, {})
                        classificacao = meta.get("classificacao", "")
                        origem_preco = meta.get("origem_preco", "")
                        preco = regime_insumo_prices[regime][uf].get(item.item_codigo, Decimal("0"))
                        if preco == 0 and uf != "SP":
                            preco_sp = regime_insumo_prices[regime]["SP"].get(item.item_codigo, Decimal("0"))
                            if preco_sp:
                                preco = preco_sp
                                preco_fallback_uf = "SP"
                    else:
                        sub = regime_costs[regime][uf].get(item.item_codigo)
                        classificacao = "COMPOSICAO AUXILIAR"
                        if sub:
                            preco = sub.custo
                        elif uf != "SP":
                            sub_sp = regime_costs[regime]["SP"].get(item.item_codigo)
                            if sub_sp:
                                preco = sub_sp.custo
                                preco_fallback_uf = "SP"

                    total = money(item.coeficiente * preco)
                    sum_items += total
                    count_items += 1
                    if (
                        preco == 0
                        and item.coeficiente != 0
                        and item.situacao not in {"SEM PREÇO", "SEM CUSTO", "EM ESTUDO"}
                    ):
                        missing_prices.append(
                            {
                                "composition_key": comp_key,
                                "codigo_composicao": comp_code,
                                "uf": uf,
                                "regime": regime,
                                "tipo_item": item.tipo_item,
                                "codigo_item": item.item_codigo,
                                "descricao_item": item.descricao,
                                "situacao": item.situacao,
                            }
                        )

                    item_key = (
                        f"SINAPI|{competencia}|{regime}|{uf}|"
                        f"{comp_code}|{item.ordem:04d}|{item.tipo_item}|{item.item_codigo}"
                    )
                    items.append(
                        {
                            "id": stable_id("item", item_key),
                            "item_key": item_key,
                            "composicao_id": comp_id,
                            "composition_key": comp_key,
                            "tipo": item_type(item.tipo_item, classificacao),
                            "codigo": item.item_codigo,
                            "descricao": item.descricao,
                            "unidade": item.unidade,
                            "coeficiente": str(item.coeficiente),
                            "preco_unitario": str(money(preco)),
                            "custo_parcial": str(total),
                            "ordem": item.ordem,
                            "linha_origem": item.linha,
                            "metadados": json.dumps(
                                {
                                    "tipo_item_sinapi": item.tipo_item,
                                    "situacao": item.situacao,
                                    "classificacao": classificacao,
                                    "origem_preco": origem_preco,
                                    "preco_fallback_uf": preco_fallback_uf,
                                },
                                ensure_ascii=False,
                                separators=(",", ":"),
                            ),
                        }
                    )

                audit.append(
                    {
                        "composition_key": comp_key,
                        "codigo": comp_code,
                        "uf": uf,
                        "regime": regime,
                        "custo_oficial": str(comp.custo),
                        "soma_itens": str(money(sum_items)),
                        "diferenca": str(money(sum_items - comp.custo)),
                        "qtd_itens": count_items,
                    }
                )

    service_rows = list(services.values())
    write_csv(
        final_out_dir / "sinapi_servicos.csv",
        service_rows,
        ["id", "service_key", "codigo_interno", "codigo_sinapi", "descricao", "unidade", "etapa", "disciplina", "situacao"],
    )
    write_csv(
        final_out_dir / "sinapi_composicoes.csv",
        compositions,
        [
            "composition_key",
            "id",
            "servico_id",
            "service_key",
            "codigo",
            "descricao",
            "unidade",
            "uf",
            "competencia",
            "regime",
            "origem",
            "custo_unitario",
            "referencia_preco",
            "grupo",
            "as_percent",
        ],
    )
    write_csv(
        final_out_dir / "sinapi_itens.csv",
        items,
        [
            "item_key",
            "id",
            "composicao_id",
            "composition_key",
            "tipo",
            "codigo",
            "descricao",
            "unidade",
            "coeficiente",
            "preco_unitario",
            "custo_parcial",
            "ordem",
            "linha_origem",
            "metadados",
        ],
    )
    write_csv(
        final_out_dir / "sinapi_auditoria_totais.csv",
        audit,
        ["composition_key", "codigo", "uf", "regime", "custo_oficial", "soma_itens", "diferenca", "qtd_itens"],
    )
    write_csv(
        final_out_dir / "sinapi_precos_ausentes.csv",
        missing_prices,
        ["composition_key", "codigo_composicao", "uf", "regime", "tipo_item", "codigo_item", "descricao_item", "situacao"],
    )

    diffs = [abs(number(row["diferenca"]) or Decimal("0")) for row in audit]
    over_cent = sum(1 for diff in diffs if diff > Decimal("0.01"))
    over_real = sum(1 for diff in diffs if diff > Decimal("1"))
    summary = {
        "fonte": "SINAPI",
        "arquivo": str(xlsx_path),
        "competencia": competencia,
        "ufs": list(UFS),
        "regimes": list(REGIMES.keys()),
        "servicos": len(service_rows),
        "composicoes": len(compositions),
        "itens": len(items),
        "precos_ausentes": len(missing_prices),
        "auditoria_diferencas_maiores_que_0_01": over_cent,
        "auditoria_diferencas_maiores_que_1_00": over_real,
        "maior_diferenca_abs": str(max(diffs) if diffs else Decimal("0")),
    }
    (final_out_dir / "sinapi_resumo.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Gera CSVs auditaveis do SINAPI para o modulo Orcamento.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_REF_XLSX, help="Caminho do SINAPI_Referencia_YYYY_MM.xlsx")
    parser.add_argument("--out-dir", type=Path, default=None, help="Pasta de saida dos CSVs")
    parser.add_argument("--ufs", nargs="+", default=list(UFS), help="UFs a extrair, por exemplo: MG RJ SP")
    args = parser.parse_args()
    main(args.xlsx, args.out_dir, tuple(args.ufs))
